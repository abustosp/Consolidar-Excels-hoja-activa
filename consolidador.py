import openpyxl
from tkinter.filedialog import askdirectory
import os
from tkinter.messagebox import showinfo

def consolidador():

    # Abrir un directorio
    directorio = askdirectory(title="Selecciona la carpeta con los archivos a consolidar")

    # Filtrar los archivos de Excel
    archivos = os.listdir(directorio)
    archivos = [archivo for archivo in archivos if archivo.endswith(".xlsx")]

    # Por cada archivo, leer la hoja y consolidarla en un solo archivo, en la columna A debe ir el nombre del archivo
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Consolidado"

    for archivo in archivos:
        libro_temp = openpyxl.load_workbook(os.path.join(directorio, archivo))
        hoja_temp = libro_temp.active
        for fila in hoja_temp.iter_rows():
            hoja.append([archivo] + [celda.value for celda in fila])
            
    libro.save("consolidado.xlsx")
    
    showinfo("Consolidador", "Se ha consolidado la información en el archivo consolidado.xlsx")